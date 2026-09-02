import argparse
import math
import warnings
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple

import json
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import seaborn as sns
from openpyxl.styles import Alignment, PatternFill
from scipy.stats import pearsonr

warnings.filterwarnings("ignore")

DEFAULT_CONFIG: Dict[str, Any] = {
	"wue_file": "data/processed/isotopes/iWUE_CO2_1901-1998.csv",
	"wue_detrended_file": "results/Paper_005/WUE/data__detrended_data.csv",
	"climate_raw_dir": "data/processed/CRU_TS_4.09",
	"climate_detrended_dir": "results/Paper_005/CRU_TS_4.09",
	"correlation_period": [1901, 1998],
	"output_dir": "results/Paper_005/correlations",
	"clim_indices": ["cld", "pet", "pre", "tmp", "rhm", "vpd"],
	"sites": ["CAN", "YAK", "TAY", "FIN", "SWE", "NOR", "ALT", "LOT", "PAK"],
	"sites_sorted": ["YAK", "TAY", "CAN", "FIN", "SWE", "NOR", "ALT", "PAK", "LOT"],
	"stat_index_full_name": {
		"tmp": "Temperature",
		"pre": "Precipitation",
		"rhm": "Relative Humidity",
		"vpd": "Vapour pressure deficit",
		"cld": "Cloud cover",
		"pet": "Potential evapotranspiration"
	},
	"year_column": "Year",
	"month_column": "Month",
	"wue_suffix": "_WUE",
	"p_threshold": 0.01,
	"font_size": 20,
	"font_family": "Times New Roman",
	"excel_raw_name": "all_correlation_results_with_trends.xlsx",
	"excel_detrended_name": "all_correlation_results_detrended.xlsx",
	"excel_first_diff_name": "all_correlation_results_first_diff.xlsx",
	"heatmap_prefix_raw": "with_trends",
	"heatmap_prefix_detrended": "detrended",
	"heatmap_prefix_first_diff": "first_diff",
	"heatmap_dpi": 300,
	"heatmap_kwargs": {
		'cmap': "seismic",
		'col_cluster': False,
		'row_cluster': False,
		'linewidths': 1,
		'linecolor': 'gray',
		'cbar_pos': (0.12, .7, .05, .18),
		'cbar_kws': {'ticks': [-.6, -.3, 0, .3, .6]},
		'vmin': -0.7, 'vmax': 0.7,
		'dendrogram_ratio': (0.2, 0.05)
	}
}


class Months(Enum):
    January   = 1
    February  = 2
    March     = 3
    April     = 4
    May       = 5
    June      = 6
    July      = 7
    August    = 8
    September = 9
    October   = 10
    November  = 11
    December  = 12

    def __str__(self):
        return self.name


def parse_args() -> argparse.Namespace:
	parser = argparse.ArgumentParser(
		description="WUE-climate monthly correlation analysis (raw and detrended)."
	)
	parser.add_argument(
		"--config",
		default="config/correlation/wue_correlation.json",
		help="Path to JSON config file."
	)
	return parser.parse_args()


def load_config(config_path: str) -> Dict[str, Any]:
	cfg = DEFAULT_CONFIG.copy()
	path = Path(config_path)

	if path.exists():
		with open(path, "r", encoding="utf-8") as f:
			user_cfg = json.load(f)
		cfg.update(user_cfg)

	cfg["output_dir"] = str(cfg["output_dir"])
	cfg["p_threshold"] = float(cfg["p_threshold"])
	cfg["heatmap_dpi"] = int(cfg["heatmap_dpi"])
	cfg["font_size"] = int(cfg["font_size"])

	period = cfg.get("correlation_period", [1901, 1998])
	if not isinstance(period, list) or len(period) != 2:
		raise ValueError("correlation_period must be a list: [start_year, end_year]")
	cfg["correlation_period"] = [int(period[0]), int(period[1])]
	return cfg


def filter_by_year_period(
	df: pd.DataFrame,
	year_column: str,
	period: Sequence[int]
) -> pd.DataFrame:
	start_year, end_year = int(period[0]), int(period[1])
	if year_column not in df.columns:
		raise ValueError(f"Column '{year_column}' not found in dataframe.")
	return df.loc[df[year_column].between(start_year, end_year)].copy()


def ensure_output_dirs(cfg: Dict[str, Any]) -> Path:
	output_dir = Path(cfg["output_dir"])
	output_dir.mkdir(parents=True, exist_ok=True)
	return output_dir


def read_climate_data(clim_indices: Sequence[str], folder: str) -> Dict[str, pd.DataFrame]:
	data: Dict[str, pd.DataFrame] = {}
	for index in clim_indices:
		path = Path(folder) / f"{index}.csv"
		data[index] = pd.read_csv(path)
	return data


def read_detrended_climate_data(clim_indices: Sequence[str], folder: str) -> Dict[str, pd.DataFrame]:
	data: Dict[str, pd.DataFrame] = {}
	for index in clim_indices:
		path = Path(folder) / f"{index}__detrended_data.csv"
		data[index] = pd.read_csv(path)
	return data


def dropna_pearsonr(x: pd.Series, y: pd.Series) -> Tuple[float, float]:
	valid = pd.DataFrame({"x": x, "y": y}).dropna()
	if len(valid) < 2:
		return np.nan, np.nan
	r, p = pearsonr(valid["x"], valid["y"])
	return float(r), float(p)


def first_order_difference_corr(x: pd.Series, y: pd.Series) -> Tuple[float, float]:
	dx = x.diff()
	dy = y.diff()
	return dropna_pearsonr(dx, dy)


def compute_monthly_correlations(
	wue_df: pd.DataFrame,
	clim_df: pd.DataFrame,
	sites_sorted: Sequence[str],
	year_column: str,
	month_column: str,
	wue_suffix: str,
	method: str = "pearson"
) -> Tuple[pd.DataFrame, pd.DataFrame]:
	rs: Dict[str, List[float]] = {site: [] for site in sites_sorted}
	ps: Dict[str, List[float]] = {site: [] for site in sites_sorted}

	for month in range(1, 13):
		month_mask = clim_df[month_column] == month
		local_clim = clim_df.loc[month_mask].copy()
		merged = wue_df.merge(local_clim, how="inner", on=year_column)

		for site in sites_sorted:
			wue_col = f"{site}{wue_suffix}"
			clim_col = site

			if wue_col not in merged.columns or clim_col not in merged.columns:
				r, p = np.nan, np.nan
			else:
				if method == "first_diff":
					r, p = first_order_difference_corr(
						merged[wue_col],
						merged[clim_col]
					)
				else:
					r, p = dropna_pearsonr(merged[wue_col], merged[clim_col])

			rs[site].append(r)
			ps[site].append(p)

	r_df = pd.DataFrame(rs, index=range(1, 13)).T
	p_df = pd.DataFrame(ps, index=range(1, 13)).T
	return r_df, p_df


def to_superscript(n: int) -> str:
	superscript_map = {
		"0": "⁰", "1": "¹", "2": "²", "3": "³", "4": "⁴",
		"5": "⁵", "6": "⁶", "7": "⁷", "8": "⁸", "9": "⁹",
		"-": "⁻"
	}
	return "".join(superscript_map.get(ch, ch) for ch in str(n))


def format_p_value(p: float) -> str:
	if pd.isna(p):
		return ""
	if p > 0.01:
		return f"p={p:.2f}"
	if p == 0:
		return "p<10" + to_superscript(-300)
	exponent = math.floor(math.log10(p))
	return "p<10" + to_superscript(exponent)


def build_combined_table(r_df: pd.DataFrame, p_df: pd.DataFrame) -> pd.DataFrame:
	combined = r_df.copy().astype(str)
	for i in range(len(r_df)):
		for col in r_df.columns:
			r_val = r_df.iloc[i][col]
			p_val = p_df.iloc[i][col]
			if pd.isna(r_val) or pd.isna(p_val):
				combined.iloc[i][col] = ""
			else:
				combined.iloc[i][col] = f"{r_val:.2f}\n({format_p_value(p_val)})"
	return combined


def save_correlation_excel(
	rs: Dict[str, pd.DataFrame],
	ps: Dict[str, pd.DataFrame],
	output_path: Path,
	p_threshold: float
) -> None:
	with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
		for index in rs.keys():
			r_data = rs[index]
			p_data = ps[index]
			combined = build_combined_table(r_data, p_data)
			combined.to_excel(writer, sheet_name=index)

			worksheet = writer.sheets[index]
			worksheet.row_dimensions[1].height = 30
			for i in range(2, len(combined) + 2):
				worksheet.row_dimensions[i].height = 35

			blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
			red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
			alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

			for i in range(len(r_data)):
				for j, col in enumerate(r_data.columns):
					r_val = r_data.iloc[i][col]
					p_val = p_data.iloc[i][col]
					cell = worksheet.cell(row=i + 2, column=j + 2)
					cell.alignment = alignment

					if not pd.isna(r_val) and not pd.isna(p_val) and p_val < p_threshold:
						if r_val > 0:
							cell.fill = red_fill
						elif r_val < 0:
							cell.fill = blue_fill


def save_heatmaps(
	rs: Dict[str, pd.DataFrame],
	ps: Dict[str, pd.DataFrame],
	cfg: Dict[str, Any],
	title_map: Dict[str, str],
	prefix: str,
	output_dir: Path
) -> None:
	heatmap_kwargs = cfg["heatmap_kwargs"]
	p_threshold = float(cfg["p_threshold"])

	for index in rs.keys():
		data = rs[index].fillna(0)
		mask = ps[index].fillna(1) > p_threshold

		hm = sns.clustermap(
			data=data,
			mask=mask,
			xticklabels=[Months(i).name for i in range(1,13)],
			**heatmap_kwargs
		)
		hm.ax_heatmap.set_title(title_map.get(index, index), fontsize=cfg["font_size"])
		hm.ax_heatmap.set_xlabel('Month', fontsize=cfg["font_size"])
		hm.ax_heatmap.set_xticklabels(hm.ax_heatmap.get_xticklabels(), rotation = 45)
		hm.ax_heatmap.yaxis.set_tick_params(labelsize=16)
		hm.ax_heatmap.set_ylabel('Site code', fontsize=cfg["font_size"])

		hm.ax_cbar.set_ylabel('Pearson R')
		hm.ax_cbar.yaxis.tick_left()
		hm.ax_cbar.yaxis.set_label_position("left")

		out_path = output_dir / f"{prefix}_{index}.png"
		plt.savefig(
			out_path,
			facecolor="white",
			transparent=False,
			dpi=cfg["heatmap_dpi"],
			bbox_inches="tight"
		)
		plt.close()


def run_correlation_set(
	wue_df: pd.DataFrame,
	climate_data: Dict[str, pd.DataFrame],
	cfg: Dict[str, Any],
	method: str = "pearson"
) -> Tuple[Dict[str, pd.DataFrame], Dict[str, pd.DataFrame]]:
	rs: Dict[str, pd.DataFrame] = {}
	ps: Dict[str, pd.DataFrame] = {}

	for index in cfg["clim_indices"]:
		r_df, p_df = compute_monthly_correlations(
			wue_df=wue_df,
			clim_df=climate_data[index],
			sites_sorted=cfg["sites_sorted"],
			year_column=cfg["year_column"],
			month_column=cfg["month_column"],
			wue_suffix=cfg["wue_suffix"],
			method=method
		)
		rs[index] = r_df
		ps[index] = p_df
	return rs, ps


def main() -> None:
	args = parse_args()
	cfg = load_config(args.config)

	plt.rcParams["font.size"] = str(cfg["font_size"])
	plt.rcParams["font.family"] = cfg["font_family"]

	output_dir = ensure_output_dirs(cfg)

	print("[INFO] Reading input datasets...")
	wue_df = pd.read_csv(cfg["wue_file"])
	wue_detrended_df = pd.read_csv(cfg["wue_detrended_file"])
	climate_raw = read_climate_data(cfg["clim_indices"], cfg["climate_raw_dir"])
	climate_detrended = read_detrended_climate_data(
		cfg["clim_indices"],
		cfg["climate_detrended_dir"]
	)

	# Keep only years inside correlation_period (inclusive).
	wue_df = filter_by_year_period(
		wue_df,
		cfg["year_column"],
		cfg["correlation_period"]
	)
	wue_detrended_df = filter_by_year_period(
		wue_detrended_df,
		cfg["year_column"],
		cfg["correlation_period"]
	)
	for index in cfg["clim_indices"]:
		climate_raw[index] = filter_by_year_period(
			climate_raw[index],
			cfg["year_column"],
			cfg["correlation_period"]
		)
		climate_detrended[index] = filter_by_year_period(
			climate_detrended[index],
			cfg["year_column"],
			cfg["correlation_period"]
		)

	print("[INFO] Computing correlations for raw climate data...")
	rs_raw, ps_raw = run_correlation_set(wue_df, climate_raw, cfg, method="pearson")

	print("[INFO] Computing correlations for detrended climate data...")
	rs_detr, ps_detr = run_correlation_set(
		wue_detrended_df,
		climate_detrended,
		cfg,
		method="pearson"
	)

	print("[INFO] Computing first-order-difference correlations from raw data...")
	rs_first_diff, ps_first_diff = run_correlation_set(
		wue_df,
		climate_raw,
		cfg,
		method="first_diff"
	)

	raw_excel = output_dir / cfg["excel_raw_name"]
	detr_excel = output_dir / cfg["excel_detrended_name"]
	first_diff_excel = output_dir / cfg["excel_first_diff_name"]

	print(f"[INFO] Saving tables: {raw_excel}")
	save_correlation_excel(rs_raw, ps_raw, raw_excel, cfg["p_threshold"])

	print(f"[INFO] Saving tables: {detr_excel}")
	save_correlation_excel(rs_detr, ps_detr, detr_excel, cfg["p_threshold"])

	print(f"[INFO] Saving tables: {first_diff_excel}")
	save_correlation_excel(
		rs_first_diff,
		ps_first_diff,
		first_diff_excel,
		cfg["p_threshold"]
	)

	title_map = cfg["stat_index_full_name"]
	print("[INFO] Saving heatmaps (raw)...")
	save_heatmaps(
		rs_raw,
		ps_raw,
		cfg,
		title_map,
		cfg["heatmap_prefix_raw"],
		output_dir
	)

	print("[INFO] Saving heatmaps (detrended)...")
	save_heatmaps(
		rs_detr,
		ps_detr,
		cfg,
		title_map,
		cfg["heatmap_prefix_detrended"],
		output_dir
	)

	print("[INFO] Saving heatmaps (first-order-difference)...")
	save_heatmaps(
		rs_first_diff,
		ps_first_diff,
		cfg,
		title_map,
		cfg["heatmap_prefix_first_diff"],
		output_dir
	)

	print("[INFO] Done.")


if __name__ == "__main__":
	main()
